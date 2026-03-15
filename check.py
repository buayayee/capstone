"""
check.py
--------
Main entry point. Given:
  1. An instructions PDF (your directive/rulebook)
  2. One or more submission PDFs

Extracts text from each and returns APPROVE / REJECT with reasons.

Usage:
  Single document:
    python check.py --instructions directives/instructions.pdf --document submissions/doc1.pdf

  Entire folder of documents:
    python check.py --instructions directives/instructions.pdf --folder submissions/

  Save results to CSV:
    python check.py --instructions directives/instructions.pdf --folder submissions/ --output results.csv
"""

import argparse
import csv
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

# Allow imports from preprocessing/ and api/
sys.path.insert(0, str(Path(__file__).parent / "preprocessing"))
sys.path.insert(0, str(Path(__file__).parent / "api"))
from extract_text import extract_text
from rule_parser import FRAUD_SIGNAL_PATTERNS, InstructionRules, parse_instructions


# ---------------------------------------------------------------------------
# Result dataclass
# ---------------------------------------------------------------------------

@dataclass
class CheckResult:
    filename: str
    verdict: str                      # "APPROVE" or "REJECT"
    reasons: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    extracted_text_preview: str = ""  # first 300 chars for reference


@dataclass
class CaseResult:
    case_name: str
    overall_verdict: str              # "APPROVE", "REJECT", or "WARNING"
    doc_results: list["CheckResult"] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Core checking logic
# ---------------------------------------------------------------------------

def _check_required_keywords(text: str, rules: InstructionRules) -> list[str]:
    """
    Universal checks every deferment document must pass.
    - 'date'  : looks for the word 'date' OR an actual date format (e.g. 5 Dec 2024)
    NOTE: Applicant name is intentionally NOT checked here — supporting documents
    are submitted with names redacted for confidentiality, and the directive itself
    never mandates that the applicant's name appear on the supporting document.
    Returns list of failed checks.
    """
    text_lower = text.lower()
    missing = []

    # Check for 'date' (literal) OR a recognisable date pattern
    date_present = (
        "date" in text_lower
        or bool(re.search(
            r"\b(\d{1,2}[\s\-/]\w+[\s\-/]\d{2,4}"   # 4 Dec 2024 or 04-Dec-24
            r"|\d{1,2}/\d{1,2}/\d{2,4}"              # 04/12/2024
            r"|\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{4}\b"  # December 2024
            r")",
            text_lower
        ))
    )
    if not date_present:
        missing.append("issue date")

    return missing


def _check_domain_category(text: str, rules: InstructionRules) -> tuple[str | None, list[str]]:
    """
    Check whether the document matches at least one deferment category.
    Returns (matched_category_name, []) on success, or (None, [warning]) on failure.
    """
    text_lower = text.lower()
    for category, keywords in rules.domain_groups.items():
        if any(kw in text_lower for kw in keywords):
            return category, []
    categories = ", ".join(rules.domain_groups.keys())
    return None, [f"Document does not clearly belong to any known deferment category ({categories})"]


def _check_institution(text: str, rules: InstructionRules) -> bool:
    """Return True if any recognised institution name is found in the document."""
    text_lower = text.lower()
    return any(inst in text_lower for inst in rules.recognised_institutions)


def _check_dates(text: str, rules: InstructionRules) -> list[str]:
    """
    Find all 4-digit years in the document and flag any outside the valid range.
    """
    flags = []
    min_year, max_year = rules.valid_year_range
    years_found = re.findall(r"\b(20\d{2})\b", text)
    for year_str in set(years_found):
        yr = int(year_str)
        if yr < min_year or yr > max_year:
            flags.append(f"Year {yr} is outside the valid range ({min_year}-{max_year})")
    return flags


def _check_fraud_signals(text: str) -> list[str]:
    """
    Scan the document text for obvious fraud signal patterns.
    """
    signals = []
    text_lower = text.lower()
    for pattern in FRAUD_SIGNAL_PATTERNS:
        match = re.search(pattern, text_lower, re.IGNORECASE)
        if match:
            signals.append(f"Suspicious pattern detected: '{match.group(0).strip()}'")
    return signals



def check_text(text: str, label: str, rules: InstructionRules) -> CheckResult:
    """
    Run all rule-based checks on already-extracted text.

    ``label`` is used as the CheckResult filename (e.g. a filename or case name).
    This is shared by both check_document() (single file) and check_case()
    (all files in a case folder concatenated into one blob).
    """
    reject_reasons: list[str] = []
    warnings: list[str] = []

    preview = text[:300].replace("\n", " ")

    # -- Step 1: Enough text to evaluate? -------------------------------------
    word_count = len(text.split())
    if word_count < 30:
        return CheckResult(
            filename=label,
            verdict="WARNING",
            reasons=[],
            warnings=[f"Too few words ({word_count}) to evaluate — blank page, scan artifact, or OCR failed"],
            extracted_text_preview=preview,
        )

    # -- Step 2: Universal required fields present? ---------------------------
    missing_keywords = _check_required_keywords(text, rules)
    if missing_keywords:
        reject_reasons.append(
            f"Missing required fields: {', '.join(missing_keywords)}"
        )

    # -- Step 3: Matches at least one deferment category? -------------------
    matched_category, category_warnings = _check_domain_category(text, rules)
    if matched_category:
        warnings.append(f"Category matched: {matched_category}")
    else:
        warnings.extend(category_warnings)

    # -- Step 4: Recognised institution? ------------------------------------
    if not _check_institution(text, rules):
        warnings.append("No recognised institution name found — verify manually")

    # -- Step 5: Date validity (warn only, don't reject for old dates) -------
    date_flags = _check_dates(text, rules)
    if date_flags:
        warnings.extend([f"NOTE: {f}" for f in date_flags])

    # -- Step 6: Fraud signals -----------------------------------------------
    fraud_signals = _check_fraud_signals(text)
    reject_reasons.extend(fraud_signals)

    verdict = "REJECT" if reject_reasons else "APPROVE"
    return CheckResult(
        filename=label,
        verdict=verdict,
        reasons=reject_reasons,
        warnings=warnings,
        extracted_text_preview=preview,
    )


def check_document(pdf_path: str, rules: InstructionRules) -> CheckResult:
    """
    Extract text from a single file and run rule-based checks on it.
    Used by single-file (--document) mode.
    """
    filename = Path(pdf_path).name
    try:
        text = extract_text(pdf_path)
    except RuntimeError as e:
        return CheckResult(filename=filename, verdict="WARNING",
                           reasons=[f"Could not process (tool missing): {e}"])
    except Exception as e:
        return CheckResult(filename=filename, verdict="WARNING",
                           reasons=[f"Could not extract text from document: {e}"])
    return check_text(text, filename, rules)


def check_document_with_ai(pdf_path: str, rules: InstructionRules,
                           ict_start: str | None = None,
                           ict_end: str | None = None) -> CheckResult:
    """
    Extract text from a single file and let the Groq LLM judge it.
    Used by single-file (--document --ai) mode.
    """
    from ai_checker import check_with_groq
    filename = Path(pdf_path).name
    try:
        text = extract_text(pdf_path)
    except RuntimeError as e:
        return CheckResult(filename=filename, verdict="WARNING",
                           reasons=[f"Could not process (tool missing): {e}"])
    except Exception as e:
        return CheckResult(filename=filename, verdict="WARNING",
                           reasons=[f"Could not extract text: {e}"])

    preview = text[:300].replace("\n", " ")
    try:
        verdict, reasons, notes = check_with_groq(
            document_text=text,
            directive_text=rules.raw_instructions,
            filename=filename,
            valid_categories=list(rules.domain_groups.keys()),
            ict_start=ict_start,
            ict_end=ict_end,
        )
    except EnvironmentError as e:
        print(f"\n[AI ERROR] {e}")
        sys.exit(1)
    except Exception as e:
        return CheckResult(filename=filename, verdict="WARNING",
                           reasons=[f"AI API error: {e}"],
                           extracted_text_preview=preview)

    return CheckResult(filename=filename, verdict=verdict,
                       reasons=reasons, warnings=notes,
                       extracted_text_preview=preview)


# ---------------------------------------------------------------------------

SUPPORTED_IMAGE_GLOBS = ["*.jpg", "*.jpeg", "*.png", "*.bmp", "*.tiff", "*.tif", "*.webp"]


def collect_supported_files(folder: Path) -> list[Path]:
    """Return all supported files (PDF + images) in a folder, sorted by name."""
    files: list[Path] = sorted(folder.glob("*.pdf"))
    for pattern in SUPPORTED_IMAGE_GLOBS:
        files += sorted(folder.glob(pattern))
    return files


def check_case(case_dir: Path, rules: InstructionRules, use_ai: bool,
               ict_start: str | None = None,
               ict_end: str | None = None) -> CaseResult:
    """
    Check an entire case folder by treating ALL supporting documents as one
    combined body of evidence.

    All files are read, their text is concatenated, and a *single* verdict is
    produced for the whole case.  This is the correct approach because:
      - A date on page 1 is evidence for the whole submission.
      - A short/blank trailing page does not dilute a multi-page letter.
      - Individual images are fragments of one document, not separate documents.

    The returned CaseResult contains one CheckResult entry whose filename is the
    case folder name (e.g. "case42").  Individual per-file extraction errors are
    surfaced as warnings, but if at least one file produced usable text the case
    is still evaluated.
    """
    files = collect_supported_files(case_dir)
    if not files:
        return CaseResult(
            case_name=case_dir.name,
            overall_verdict="WARNING",
            doc_results=[CheckResult(
                filename=case_dir.name,
                verdict="WARNING",
                reasons=[],
                warnings=["Case folder is empty — no supported files found"],
            )],
        )

    # ── Extract + concatenate all pages ──────────────────────────────────────
    parts: list[str] = []
    extraction_warnings: list[str] = []

    for f in files:
        try:
            t = extract_text(str(f))
            if t.strip():
                parts.append(t)
        except RuntimeError as e:
            extraction_warnings.append(f"{f.name}: tool missing — {e}")
        except Exception as e:
            extraction_warnings.append(f"{f.name}: extraction failed — {e}")

    combined_text = "\n\n".join(parts)

    # ── Evaluate the combined text ────────────────────────────────────────────
    if use_ai:
        from ai_checker import check_with_groq  # noqa: F401 — Bedrock-backed
        preview = combined_text[:300].replace("\n", " ")
        try:
            verdict, reasons, notes = check_with_groq(
                document_text=combined_text,
                directive_text=rules.raw_instructions,
                filename=case_dir.name,
                valid_categories=list(rules.domain_groups.keys()),
                ict_start=ict_start,
                ict_end=ict_end,
            )
        except EnvironmentError as e:
            print(f"\n[AI ERROR] {e}")
            sys.exit(1)
        except Exception as e:
            verdict, reasons, notes = "WARNING", [f"AI API error: {e}"], []
        case_result = CheckResult(
            filename=case_dir.name,
            verdict=verdict,
            reasons=reasons,
            warnings=notes + extraction_warnings,
            extracted_text_preview=preview,
        )
    else:
        case_result = check_text(combined_text, case_dir.name, rules)
        case_result.warnings.extend(extraction_warnings)

    # Show actual document filenames instead of the folder name
    case_result.filename = ", ".join(f.name for f in files)

    return CaseResult(
        case_name=case_dir.name,
        overall_verdict=case_result.verdict,
        doc_results=[case_result],
    )


def print_case_result(case: CaseResult):
    if case.overall_verdict == "APPROVE":
        verdict_display = "[APPROVE]"
    elif case.overall_verdict == "REJECT":
        verdict_display = "[REJECT] "
    else:
        verdict_display = "[WARNING] could not fully process"

    print("\n" + "="*60)
    print(f"  Case    : {case.case_name}  |  Verdict: {verdict_display}")
    # One combined result per case
    for result in case.doc_results:
        print_result(result, indent="    ")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def print_result(result: CheckResult, indent: str = ""):
    if result.verdict == "APPROVE":
        verdict_display = "[APPROVE]"
    elif result.verdict == "REJECT":
        verdict_display = "[REJECT] "
    else:
        verdict_display = "[WARNING] could not process"
    print(f"\n{indent}" + "-"*56)
    print(f"{indent}  File    : {result.filename}")
    print(f"{indent}  Verdict : {verdict_display}")
    if result.reasons:
        print(f"{indent}  Reasons :")
        for r in result.reasons:
            print(f"{indent}    - {r}")
    if result.warnings:
        print(f"{indent}  Notes:")
        for w in result.warnings:
            print(f"{indent}    ! {w}")
    print(f"{indent}  Preview : {result.extracted_text_preview[:120]}...")


def save_results_to_csv(cases: list[CaseResult], output_path: str):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["case", "overall_verdict", "filename", "doc_verdict", "reasons", "notes"])
        for case in cases:
            for doc in case.doc_results:
                writer.writerow([
                    case.case_name,
                    case.overall_verdict,
                    doc.filename,
                    doc.verdict,
                    " | ".join(doc.reasons),
                    " | ".join(doc.warnings),
                ])
    print(f"\nResults saved to: {output_path}")


def save_results_to_excel(cases: list[CaseResult], output_path: str):
    """Save results to a formatted Excel workbook (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        print("[WARNING] openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    GREY   = PatternFill("solid", fgColor="D9D9D9")
    NAVY   = PatternFill("solid", fgColor="1F4E79")
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Case", "Overall Verdict", "Filename", "Doc Verdict", "Reject Reasons", "Notes"]
    ws.append(headers)
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.fill = NAVY
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    row_idx = 2
    for case in cases:
        overall = case.overall_verdict
        n_docs  = len(case.doc_results)
        first   = row_idx
        for doc in case.doc_results:
            ws.cell(row=row_idx, column=1).value = case.case_name
            ws.cell(row=row_idx, column=2).value = overall
            ws.cell(row=row_idx, column=3).value = doc.filename
            ws.cell(row=row_idx, column=4).value = doc.verdict
            ws.cell(row=row_idx, column=5).value = "\n".join(doc.reasons)
            ws.cell(row=row_idx, column=6).value = "\n".join(doc.warnings)
            v_fill = GREEN if overall == "APPROVE" else (RED if overall == "REJECT" else YELLOW)
            d_fill = GREEN if doc.verdict == "APPROVE" else (RED if doc.verdict == "REJECT" else YELLOW)
            wrap   = Alignment(wrap_text=True, vertical="top")
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = border
                ws.cell(row=row_idx, column=col).alignment = wrap
            ws.cell(row=row_idx, column=2).fill = v_fill
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_idx, column=4).fill = d_fill
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1
        if n_docs > 1:
            ws.merge_cells(start_row=first, start_column=1, end_row=row_idx-1, end_column=1)
            ws.merge_cells(start_row=first, start_column=2, end_row=row_idx-1, end_column=2)
            ws.cell(row=first, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=first, column=2).alignment = Alignment(horizontal="center", vertical="center")

    for col, width in zip("ABCDEF", [14, 16, 38, 14, 55, 55]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    approve = sum(1 for c in cases if c.overall_verdict == "APPROVE")
    reject  = sum(1 for c in cases if c.overall_verdict == "REJECT")
    warn    = sum(1 for c in cases if c.overall_verdict == "WARNING")
    for r, (label, val, fill) in enumerate([
        ("Metric", "Count", GREY), ("Total Cases", len(cases), GREY),
        ("APPROVE", approve, GREEN), ("REJECT", reject, RED), ("WARNING", warn, YELLOW),
    ], 1):
        ws2.cell(r, 1, label); ws2.cell(r, 2, val)
        for col in range(1, 3):
            ws2.cell(r, col).fill   = fill
            ws2.cell(r, col).border = border
            ws2.cell(r, col).alignment = Alignment(horizontal="center")
    ws2.cell(1,1).font = Font(bold=True); ws2.cell(1,2).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    print(f"\nExcel results saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Check deferment supporting PDFs against the instructions directive."
    )
    parser.add_argument(
        "--instructions", required=True,
        help="Path to the instructions/directive file (.docx or .pdf)"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--document",
        help="Path to a single submission file to check (.pdf, .jpg, .png, etc.)"
    )
    group.add_argument(
        "--folder",
        help="Path to a folder of submission files to check in bulk (.pdf, .jpg, .png, etc.)"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="(Optional) Save results to this CSV file path"
    )
    parser.add_argument(
        "--ai",
        action="store_true",
        help="Use Groq LLM to judge documents (requires GROQ_API_KEY env variable)"
    )
    parser.add_argument(
        "--ict-start",
        default=None,
        metavar="DATE",
        help="ICT/training start date (e.g. '01 Jan 2023'). Used by AI mode for temporal checks."
    )
    parser.add_argument(
        "--ict-end",
        default=None,
        metavar="DATE",
        help="ICT/training end date (e.g. '13 Jan 2023'). Used by AI mode for temporal checks."
    )
    args = parser.parse_args()

    # ── Parse instructions ────────────────────────────────────────────────────
    rules = parse_instructions(args.instructions)

    # ── Mode header ──────────────────────────────────────────────────────────
    if args.ai:
        print("[Mode] AI-powered checks via Amazon Bedrock (Claude 3 Haiku)")
    else:
        print("[Mode] Rule-based checks (use --ai to enable Groq AI)")

    # ── Single document ───────────────────────────────────────────────────────
    if args.document:
        result = (
            check_document_with_ai(args.document, rules,
                                   ict_start=args.ict_start, ict_end=args.ict_end)
            if args.ai
            else check_document(args.document, rules)
        )
        # Wrap in a single-doc case so output is consistent
        case = CaseResult(
            case_name=Path(args.document).stem,
            overall_verdict=result.verdict,
            doc_results=[result],
        )
        print_case_result(case)
        if args.output:
            save_results_to_csv([case], args.output)
        return

    # ── Folder mode ───────────────────────────────────────────────────────────
    folder = Path(args.folder)

    # Detect whether submissions are organised into case subfolders
    case_dirs = sorted(
        [d for d in folder.iterdir() if d.is_dir()],
        key=lambda d: (
            int(d.name[4:]) if d.name.startswith("case") and d.name[4:].isdigit() else float("inf")
        )
    )

    if case_dirs:
        # ── Case-folder mode: submissions/case1/, case2/, ... ─────────────────
        print(f"\nFound {len(case_dirs)} case folder(s) in {folder}")
        cases: list[CaseResult] = []
        approved = rejected = warnings = 0

        for case_dir in case_dirs:
            case = check_case(case_dir, rules, use_ai=args.ai,
                             ict_start=args.ict_start, ict_end=args.ict_end)
            print_case_result(case)
            cases.append(case)
            if case.overall_verdict == "APPROVE":
                approved += 1
            elif case.overall_verdict == "REJECT":
                rejected += 1
            else:
                warnings += 1

        print("\n" + "="*60)
        print(f"  SUMMARY: {len(cases)} case(s) checked")
        print(f"  [APPROVE]  : {approved}")
        print(f"  [REJECT]   : {rejected}")
        print(f"  [WARNING]  : {warnings}")
        print("="*60)

        if args.output:
            save_results_to_csv(cases, args.output)

    else:
        # ── Flat mode (backwards compat): submissions/*.pdf ───────────────────
        files = collect_supported_files(folder)
        if not files:
            print(f"[ERROR] No supported files or case folders found in: {folder}")
            sys.exit(1)
        print(f"\nFound {len(files)} file(s) in {folder} (flat mode)")

        cases = []
        approved = rejected = warnings = 0
        for f in files:
            result = (
                check_document_with_ai(str(f), rules,
                                       ict_start=args.ict_start,
                                       ict_end=args.ict_end)
                if args.ai
                else check_document(str(f), rules)
            )
            case = CaseResult(
                case_name=f.stem,
                overall_verdict=result.verdict,
                doc_results=[result],
            )
            print_case_result(case)
            cases.append(case)
            if result.verdict == "APPROVE":
                approved += 1
            elif result.verdict == "REJECT":
                rejected += 1
            else:
                warnings += 1

        print("\n" + "="*60)
        print(f"  SUMMARY: {len(cases)} document(s) checked")
        print(f"  [APPROVE]  : {approved}")
        print(f"  [REJECT]   : {rejected}")
        print(f"  [WARNING]  : {warnings}")
        print("="*60)

        if args.output:
            save_results_to_csv(cases, args.output)


if __name__ == "__main__":
    main()
